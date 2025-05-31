# Usage
# python host_matcher.py ^
#   --jira-url http://jira.internal ^
#   --jql "project = NETSEC AND issuetype = Feature" ^
#   --trigger-pattern "server names.*?encrypting is hosted on:" ^
#   --file-b valid_hosts.xlsx

import re
import csv
import unicodedata
import argparse
import requests
import pickle
from openpyxl import load_workbook
from collections import defaultdict

# --- Cleaning Function ---
def clean_string(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("â€“", "-").replace("â€”", "-")
    s = s.replace("â€˜", "'").replace("â€™", "'")
    s = s.replace("â€œ", '"').replace("â€", '"')
    s = s.replace("â€¦", "...").replace("Â", "").replace("â", "")
    s = s.replace('\u00A0', ' ').replace('\u200B', '').replace('\uFEFF', '')
    s = re.sub(r'[\u2013\u2014]', '-', s)
    s = re.sub(r'[\u2018\u2019\u201A]', "'", s)
    s = re.sub(r'[\u201C\u201D\u201E]', '"', s)
    s = re.sub(r'[^\x00-\x7F]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()

# --- Jira Fetch with Pagination ---
def fetch_jira_hostnames(jira_url, token, jql, trigger_pattern):
    results = []
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }

    max_results = 100
    start_at = 0
    total = None

    while total is None or start_at < total:
        params = {
            "jql": jql,
            "fields": "key,description",
            "maxResults": max_results,
            "startAt": start_at
        }

        url = f"{jira_url}/rest/api/2/search"
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()

        total = data.get("total", 0)
        issues = data.get("issues", [])
        print(f"🔄 Retrieved {len(issues)} issues (startAt={start_at})...", flush=True)

        for issue in issues:
            key = issue["key"]
            desc = issue["fields"].get("description", "")
            cleaned_text = clean_string(desc)
            match = re.search(trigger_pattern, cleaned_text)
            if match:
                after = cleaned_text[match.end():].strip()
                first_line = after.splitlines()[0]
                servers = first_line.split(',')
                for server in servers:
                    hostname = clean_string(server).split('.')[0]
                    if hostname:
                        results.append((key, hostname))

        start_at += max_results

    return results

# --- Main CLI Execution ---
def main():
    parser = argparse.ArgumentParser(description="Extract and match hostnames from Jira or Excel.")
    parser.add_argument("--file-a", help="Excel file with source data")
    parser.add_argument("--file-b", required=True, help="Excel file with valid hostnames (column to match against)")
    parser.add_argument("--trigger-pattern", required=True, help="Regex trigger pattern before the host list")
    parser.add_argument("--extract-column", type=int, default=7, help="Column in file-a with hostname string")
    parser.add_argument("--id-column", type=int, default=2, help="Column in file-a with ID value")
    parser.add_argument("--reference-column", type=int, default=25, help="Column in file-b with valid hostnames")
    parser.add_argument("--output-prefix", default="host_check", help="Prefix for output CSVs")
    parser.add_argument("--jira-url", help="Jira base URL")
    parser.add_argument("--jql", help="JQL query to pull Jira issues")

    args = parser.parse_args()

    # --- Step 1: Get extracted hostnames (from Jira or Excel A) ---
    extracted = []

    if args.jira_url and args.jql:
        print("🔗 Using Jira as input source...")
        try:
            # Load token from cache
            with open("jira_token_cache.pkl", "rb") as f:
                token = pickle.load(f)
            extracted = fetch_jira_hostnames(
                args.jira_url, token, args.jql, args.trigger_pattern
            )
            print(f"✅ Retrieved {len(extracted)} hostnames from Jira issues.")
        except Exception as e:
            print(f"❌ Jira fetch failed: {e}")
            return
    elif args.file_a:
        print("📂 Using Excel file A as input source...")
        wb_a = load_workbook(args.file_a, data_only=True)
        for sheet in wb_a.worksheets:
            for i in range(1, sheet.max_row + 1):
                text_val = sheet.cell(row=i, column=args.extract_column).value
                id_val = sheet.cell(row=i, column=args.id_column).value
                cleaned_text = clean_string(text_val)
                id_clean = clean_string(id_val)
                match = re.search(args.trigger_pattern, cleaned_text)
                if match:
                    after = cleaned_text[match.end():].strip()
                    first_line = after.splitlines()[0]
                    raw_candidates = first_line.split(',')
                    for server in raw_candidates:
                        hostname = clean_string(server).split('.')[0]
                        if hostname:
                            extracted.append((id_clean, hostname))
    else:
        print("❌ Provide either Jira parameters or --file-a.")
        return

    # --- Step 2: Load valid hosts from File B ---
    valid_hosts = set()
    wb_b = load_workbook(args.file_b, data_only=True)
    for sheet in wb_b.worksheets:
        for i in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=args.reference_column).value
            hostname = clean_string(cell).split('.')[0]
            if hostname:
                valid_hosts.add(hostname)

    # --- Step 3: Compare results ---
    output_rows = []
    for id_val, hostname in extracted:
        status = "Matched" if hostname in valid_hosts else "Unmatched"
        output_rows.append((id_val, hostname, status))

    # --- Step 4: Write details CSV ---
    detail_csv = f"{args.output_prefix}_details.csv"
    with open(detail_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'Hostname', 'Status'])
        writer.writerows(output_rows)
    print(f"✅ Detailed output saved to {detail_csv}")

    # --- Step 5: Summary by ID ---
    summary = defaultdict(lambda: {'Total': 0, 'Matched': 0, 'Unmatched': 0})
    for id_val, _, status in output_rows:
        key = id_val or "(blank)"
        summary[key]['Total'] += 1
        if status == "Matched":
            summary[key]['Matched'] += 1
        else:
            summary[key]['Unmatched'] += 1

    summary_csv = f"{args.output_prefix}_summary.csv"
    with open(summary_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'Total Hosts', 'Matched', 'Unmatched'])
        for id_val, counts in summary.items():
            writer.writerow([
                id_val,
                counts['Total'],
                counts['Matched'],
                counts['Unmatched']
            ])
    print(f"📊 Summary output saved to {summary_csv}")

if __name__ == "__main__":
    main()
