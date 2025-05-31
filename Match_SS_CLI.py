# Match host names regex'd from one column of a SS to a master lookup in a col of another SS
# python host_matcher.py --file-a spreadsheet_a.xlsx --file-b spreadsheet_b.xlsx ^
#  --trigger-pattern "server names.*?encrypting is hosted on:" ^
#  --extract-column 7 --id-column 2 --reference-column 25 --output-prefix host_output

import re
import csv
import unicodedata
import argparse
from openpyxl import load_workbook
from collections import defaultdict

# --- Cleaning Function ---
def clean_string(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("â€“", "-").replace("â€”", "-")
    s = s.replace("â€˜", "'").replace("â€™", "'")
    s = s.replace("â€œ", '"').replace("â€", '"')
    s = s.replace("â€¦", "...").replace("Â", "").replace("â", "")
    s = s.replace('\u00A0', ' ').replace('\u200B', '').replace('\uFEFF', '')
    s = re.sub(r'[\u2013\u2014]', '-', s)
    s = re.sub(r'[\u2018\u2019\u201A]', "'", s)
    s = re.sub(r'[\u201C\u201D\u201E]', '"', s)
    s = re.sub(r'[^\x00-\x7F]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()

# --- Main Execution ---
def main():
    parser = argparse.ArgumentParser(description="Extract and compare hostnames from Excel files.")
    parser.add_argument("--file-a", required=True, help="Input spreadsheet A (with trigger + hostnames)")
    parser.add_argument("--file-b", required=True, help="Reference spreadsheet B (with valid hostnames)")
    parser.add_argument("--trigger-pattern", required=True, help="Regex pattern that precedes the host list")
    parser.add_argument("--extract-column", type=int, default=7, help="Column number to extract host list from (default 7 for G)")
    parser.add_argument("--id-column", type=int, default=2, help="Column number for ID/grouping (default 2 for B)")
    parser.add_argument("--reference-column", type=int, default=25, help="Column with valid hostnames (default 25 for Y)")
    parser.add_argument("--output-prefix", default="host_check", help="Prefix for output files")

    args = parser.parse_args()

    extracted = []  # (id_val, hostname)

    wb_a = load_workbook(args.file_a, data_only=True)
    for sheet in wb_a.worksheets:
        for i in range(1, sheet.max_row + 1):
            text_val = sheet.cell(row=i, column=args.extract_column).value
            id_val = sheet.cell(row=i, column=args.id_column).value
            cleaned_text = clean_string(text_val)
            id_clean = clean_string(id_val)

            match = re.search(args.trigger_pattern, cleaned_text)
            if match:
                after = cleaned_text[match.end():].strip()
                first_line = after.splitlines()[0]
                raw_candidates = first_line.split(',')

                for server in raw_candidates:
                    hostname = clean_string(server).split('.')[0]
                    if hostname:
                        extracted.append((id_clean, hostname))

    valid_hosts = set()
    wb_b = load_workbook(args.file_b, data_only=True)
    for sheet in wb_b.worksheets:
        for i in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=args.reference_column).value
            hostname = clean_string(cell).split('.')[0]
            if hostname:
                valid_hosts.add(hostname)

    output_rows = []
    for id_val, hostname in extracted:
        status = "Matched" if hostname in valid_hosts else "Unmatched"
        output_rows.append((id_val, hostname, status))

    # --- Write detailed output
    detail_csv = f"{args.output_prefix}_details.csv"
    with open(detail_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'Hostname', 'Status'])
        writer.writerows(output_rows)
    print(f"✅ Detailed results written to {detail_csv}")

    # --- Write summary output
    summary = defaultdict(lambda: {'Total': 0, 'Matched': 0, 'Unmatched': 0})
    for id_val, hostname, status in output_rows:
        key = id_val or "(blank)"
        summary[key]['Total'] += 1
        if status == "Matched":
            summary[key]['Matched'] += 1
        else:
            summary[key]['Unmatched'] += 1

    summary_csv = f"{args.output_prefix}_summary.csv"
    with open(summary_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'Total Hosts', 'Matched', 'Unmatched'])
        for id_val, counts in summary.items():
            writer.writerow([
                id_val,
                counts['Total'],
                counts['Matched'],
                counts['Unmatched']
            ])
    print(f"📊 Summary written to {summary_csv}")

if __name__ == "__main__":
    main()
