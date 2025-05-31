import re
import csv
import unicodedata
from openpyxl import load_workbook
from collections import defaultdict

# --- Config ---
file_a = 'spreadsheet_a.xlsx'
file_b = 'spreadsheet_b.xlsx'
output_csv = 'matched_hostnames.csv'
trigger = "server names that the file or folder that needs encrypting is hosted on:"

# --- Cleaning Function ---
def clean_string(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("â€“", "-").replace("â€”", "-")
    s = s.replace("â€˜", "'").replace("â€™", "'")
    s = s.replace("â€œ", '"').replace("â€", '"')
    s = s.replace("â€¦", "...").replace("Â", "").replace("â", "")
    s = s.replace('\u00A0', ' ').replace('\u200B', '').replace('\uFEFF', '')
    s = re.sub(r'[\u2013\u2014]', '-', s)
    s = re.sub(r'[\u2018\u2019\u201A]', "'", s)
    s = re.sub(r'[\u201C\u201D\u201E]', '"', s)
    s = re.sub(r'[^\x00-\x7F]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()

# --- Step 1: Extract from File A ---
extracted = []  # (col_b_val, hostname)

wb_a = load_workbook(file_a, data_only=True)
for sheet in wb_a.worksheets:
    for i in range(1, sheet.max_row + 1):
        col_g_val = sheet.cell(row=i, column=7).value  # Column G
        col_b_val = sheet.cell(row=i, column=2).value  # Column B
        col_g_val_clean = clean_string(col_g_val)
        col_b_val_clean = clean_string(col_b_val)

        if trigger in col_g_val_clean:
            after = col_g_val_clean.split(trigger, 1)[1].strip()
            first_line = after.splitlines()[0]
            raw_candidates = first_line.split(',')
            for server in raw_candidates:
                hostname = clean_string(server).split('.')[0]
                if hostname:
                    extracted.append((col_b_val_clean, hostname))

# --- Step 2: Load Valid Hostnames from File B (Column Y = 25) ---
valid_hosts = set()
wb_b = load_workbook(file_b, data_only=True)
for sheet in wb_b.worksheets:
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=i, column=25).value
        if isinstance(cell, str):
            hostname = clean_string(cell).split('.')[0]
            if hostname:
                valid_hosts.add(hostname)

# --- Step 3: Compare + Prepare Full Results ---
output_rows = []
for col_b_val, hostname in extracted:
    status = "Matched" if hostname in valid_hosts else "Unmatched"
    output_rows.append((col_b_val, hostname, status))

# --- Step 4: Write Detailed CSV ---
with open(output_csv, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Column B Value', 'Hostname', 'Status'])
    writer.writerows(output_rows)

print(f"✅ Detailed results written to {output_csv}")

# --- Step 5: Build and Write Summary CSV ---
summary = defaultdict(lambda: {'Total': 0, 'Matched': 0, 'Unmatched': 0})
for col_b_val, hostname, status in output_rows:
    key = col_b_val or "(blank)"
    summary[key]['Total'] += 1
    if status == "Matched":
        summary[key]['Matched'] += 1
    else:
        summary[key]['Unmatched'] += 1

summary_csv = output_csv.replace('.csv', '_summary.csv')
with open(summary_csv, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Column B Value', 'Total Hosts', 'Matched', 'Unmatched'])
    for col_b_val, counts in summary.items():
        writer.writerow([
            col_b_val,
            counts['Total'],
            counts['Matched'],
            counts['Unmatched']
        ])

print(f"📊 Summary written to {summary_csv}")
