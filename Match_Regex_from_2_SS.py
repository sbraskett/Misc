import re
import csv
import unicodedata
from openpyxl import load_workbook
from collections import defaultdict

# --- Configurable Parameters ---
file_a = 'spreadsheet_a.xlsx'
file_b = 'spreadsheet_b.xlsx'
trigger_pattern = r"server names that the file or folder that needs encrypting is hosted on:"  # regex
extract_column = 7   # e.g., G
id_column = 2        # e.g., B
reference_column = 25  # e.g., Y
output_prefix = "host_check"

# --- Cleaning Function ---
def clean_string(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("â€“", "-").replace("â€”", "-")
    s = s.replace("â€˜", "'").replace("â€™", "'")
    s = s.replace("â€œ", '"').replace("â€", '"')
    s = s.replace("â€¦", "...").replace("Â", "").replace("â", "")
    s = s.replace('\u00A0', ' ').replace('\u200B', '').replace('\uFEFF', '')
    s = re.sub(r'[\u2013\u2014]', '-', s)
    s = re.sub(r'[\u2018\u2019\u201A]', "'", s)
    s = re.sub(r'[\u201C\u201D\u201E]', '"', s)
    s = re.sub(r'[^\x00-\x7F]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()

# --- Step 1: Extract from File A ---
extracted = []  # (id_value, hostname)

wb_a = load_workbook(file_a, data_only=True)
for sheet in wb_a.worksheets:
    for i in range(1, sheet.max_row + 1):
        text_val = sheet.cell(row=i, column=extract_column).value
        id_val = sheet.cell(row=i, column=id_column).value
        cleaned_text = clean_string(text_val)
        id_clean = clean_string(id_val)

        match = re.search(trigger_pattern, cleaned_text)
        if match:
            # Extract everything after the match
            after = cleaned_text[match.end():].strip()
            first_line = after.splitlines()[0]
            raw_candidates = first_line.split(',')

            for server in raw_candidates:
                hostname = clean_string(server).split('.')[0]
                if hostname:
                    extracted.append((id_clean, hostname))

# --- Step 2: Load Valid Hostnames from File B ---
valid_hosts = set()
wb_b = load_workbook(file_b, data_only=True)
for sheet in wb_b.worksheets:
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=i, column=reference_column).value
        hostname = clean_string(cell).split('.')[0]
        if hostname:
            valid_hosts.add(hostname)

# --- Step 3: Compare + Prepare Results ---
output_rows = []
for id_val, hostname in extracted:
    status = "Matched" if hostname in valid_hosts else "Unmatched"
    output_rows.append((id_val, hostname, status))

# --- Step 4: Write Detailed CSV ---
output_csv = f"{output_prefix}_details.csv"
with open(output_csv, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['ID', 'Hostname', 'Status'])
    writer.writerows(output_rows)

print(f"✅ Detailed results written to {output_csv}")

# --- Step 5: Write Summary CSV ---
summary = defaultdict(lambda: {'Total': 0, 'Matched': 0, 'Unmatched': 0})
for id_val, hostname, status in output_rows:
    key = id_val or "(blank)"
    summary[key]['Total'] += 1
    if status == "Matched":
        summary[key]['Matched'] += 1
    else:
        summary[key]['Unmatched'] += 1

summary_csv = f"{output_prefix}_summary.csv"
with open(summary_csv, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['ID', 'Total Hosts', 'Matched', 'Unmatched'])
    for id_val, counts in summary.items():
        writer.writerow([
            id_val,
            counts['Total'],
            counts['Matched'],
            counts['Unmatched']
        ])

print(f"📊 Summary written to {summary_csv}")
